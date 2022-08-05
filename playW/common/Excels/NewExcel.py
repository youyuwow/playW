# coding:utf8
import os, openpyxl
from shutil import copyfile
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection


class Reader:
    """
        用来读取Excel文件内容
        只支持xlsx格式excel文件读取
        pip install -U openpyxl
        https://openpyxl.readthedocs.io/en/default/styles.html
    """

    def __init__(self):
        # 整个excel工作簿缓存
        self.workbook = None
        # 当前工作sheet
        self.sheet = None
        # 当前sheet的行数
        self.rows = 0
        # 当前读取到的行数
        self.r = 0

    # 打开excel
    def open_excel(self, srcfile):
        # 如果打开的文件不存在，就报错
        if not os.path.isfile(srcfile):
            print("%s not exist!" % (srcfile))
            return

        # 设置读取excel使用utf8编码
        openpyxl.Workbook.encoding = "utf8"
        # 读取excel内容到缓存workbook
        self.workbook = openpyxl.load_workbook(filename=srcfile)
        # 设置默认选取第一个sheet页面
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 设置rows为当前sheet的行数
        self.rows = self.sheet.max_row
        # 设置默认读取为第一行
        self.r = 0
        return

    # 获取sheet页面
    def get_sheets(self):
        # 获取所有sheet的名字，并返回为一个列表
        sheets = self.workbook.sheetnames
        # print(sheets)
        return sheets

    # 切换sheet页面
    def set_sheet(self, name):
        # 通过sheet名字，切换sheet页面
        self.sheet = self.workbook[name]
        self.rows = self.sheet.max_row
        self.r = 0
        return

    # 逐行读取
    def readline(self):
        # 存sheet里面所有的行
        lines = []
        # 遍历sheet里面所有的行
        for row in self.sheet.rows:
            # 存一行的单元格
            line = []
            # 取出所有的单元格
            for cell in row:
                if cell.value is None:
                    line.append('')
                else:
                    line.append(cell.value)

            lines.append(line)

        return lines


class Writer:
    """
        用来复制写入excel表格内容
        只支持xlsx格式excel文件读取
        pip install -U openpyxl
        https://openpyxl.readthedocs.io/en/default/styles.html
    """

    def __init__(self):
        # 读取需要复制的excel
        self.workbook = None
        # 拷贝的工作空间
        self.wb = None
        # 当前工作的sheet页
        self.sheet = None
        # 记录生成的文件，用来保存
        self.df = None
        # 记录写入的行
        self.row = 0
        # 记录写入的列
        self.clo = 0

    # 复制并打开excel
    def copy_open(self, srcfile, dstfile):
        # 判断要复制的文件是否存在
        if not os.path.isfile(srcfile):
            print(srcfile + " not exist!")
            return

        # 判断要新建的文档是否存在，存在则提示
        if os.path.isfile(dstfile):
            print(dstfile + " file already exist!")

        # 记录要保存的文件
        self.df = dstfile
        # 读取excel到缓存
        # formatting_info带格式的复制
        self.workbook = openpyxl.load_workbook(filename=srcfile)
        # 拷贝，也在内存里面
        copyfile(srcfile, dstfile)
        # 打开复制后的excel文件
        self.wb = openpyxl.load_workbook(filename=dstfile)
        return

    # 获取sheet页面
    def get_sheets(self):
        # 获取所有sheet的名字，并返回为一个列表
        sheets = self.workbook.sheetnames
        return sheets

    # 切换sheet页面
    def set_sheet(self, name):
        # 通过sheet名字，切换sheet页面
        self.sheet = self.wb[name]
        return

    # 写入指定单元格，保留原格式
    def write(self, r, c, value, color=None):
        """
        :param r: 行
        :param c: 列
        :param value: 要写入的字符串
        :param color: 0，黑色；1，白色；2，红色；3，绿色；4，蓝色；5，黄色
        :return:
        """
        d = self.sheet.cell(row=r + 1, column=c + 1, value=value)

        if color:
            if color == 0:
                color = "FF000000"
            elif color == 1:
                color = "FFFFFFFF"
            elif color == 2:
                color = "FFFF0000"
            elif color == 3:
                color = "FF00FF00"
            elif color == 4:
                color = "FF0000FF"
            elif color == 5:
                color = "FFFFFF00"
            else:
                color = "FF000000"

        font = Font(name='Arial',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color=color)

        d.font = font

        # fill = PatternFill(fill_type=None,
        #                    start_color='FFFFFFFF',
        #                    end_color='FF000000')
        # border = Border(left=Side(border_style=None,
        #                           color='FF000000'),
        #                 right=Side(border_style=None,
        #                            color='FF000000'),
        #                 top=Side(border_style=None,
        #                          color='FF000000'),
        #                 bottom=Side(border_style=None,
        #                             color='FF000000'),
        #                 diagonal=Side(border_style=None,
        #                               color='FF000000'),
        #                 diagonal_direction=0,
        #                 outline=Side(border_style=None,
        #                              color='FF000000'),
        #                 vertical=Side(border_style=None,
        #                               color='FF000000'),
        #                 horizontal=Side(border_style=None,
        #                                 color='FF000000'))
        # alignment = Alignment(horizontal='general',
        #                       vertical='bottom',
        #                       text_rotation=0,
        #                       wrap_text=False,
        #                       shrink_to_fit=False,
        #                       indent=0)
        # number_format = 'General'
        # protection = Protection(locked=True,
        #                         hidden=False)

        return

    # 保存
    def save_close(self):
        # 保存复制后的文件到硬盘
        self.wb.save(self.df)
        return


if __name__ == '__main__':
    reader = Reader()
    # 打开一个excel
    reader.open_excel('../../lib/cases/电商项目用例.xlsx')
    # 获取所有sheet
    sheets = reader.get_sheets()
    for sheet in sheets:
        # 设置读取的sheet页面
        reader.set_sheet(sheet)
        # 读取当前sheet的所有行
        lines = reader.readline()
        print(lines)

    writer = Writer()
    writer.copy_open('../../lib/cases/电商项目用例.xlsx','../../lib/cases/result-电商项目用例.xlsx')

    # 获取所有sheet
    sheets = writer.get_sheets()
    writer.set_sheet(sheets[0])
    # 在原格式上写入内容
    writer.write(1,1,'Will')
    # 更改颜色写入
    writer.write(1,2,'hello',color=2)
    writer.save_close()